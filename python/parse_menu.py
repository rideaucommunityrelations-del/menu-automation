"""
Parses the chef's raw weekly menu excel into 7 day-dicts ready for render_menu.render_day().
Robust to the messy label spacing in the raw file (matches by substring, not exact position).
"""
import re
import openpyxl
from datetime import datetime

DAY_COLS = ["B", "C", "D", "E", "F", "G", "H"]  # Monday -> Sunday, fixed order in the raw file
DAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

LABELS = {
    "daily_soup": "daily soup",
    "soup_of_week": "soup of the week",
    "sandwich_of_day": "sandwich of the day",
    "lunch_entree": "lunch entr",       # matches "Lunch Entree" / "Lunch Entrée"
    "appetizer": "appetizer",
    "dinner_entree": "dinner entr",     # matches both Dinner Entree rows; grab 1st then next occurrence
    "dinner_dessert": "dinner dessert",
}


def _clean(s):
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip()


def parse_weekly_menu(xlsx_path, year=None):
    """
    Returns a list of 7 dicts (Monday..Sunday), each shaped for render_menu.render_day().
    `year` defaults to the current year; pass explicitly if running near a year boundary.
    """
    if year is None:
        year = datetime.now().year

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    # Build a list of (cleaned column-A label, row number) - keep all rows, including duplicate labels
    label_rows = []
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if cell.value is None:
            continue
        label_rows.append((_clean(cell.value).lower(), cell.row))

    def find_row(substring, occurrence=1):
        matches = sorted(r for lbl, r in label_rows if substring in lbl)
        if len(matches) < occurrence:
            raise ValueError(f"Could not find row for label containing '{substring}' "
                              f"(occurrence {occurrence}). Found labels: {[l for l,_ in label_rows]}")
        return matches[occurrence - 1]

    # Find the header row: the row where column B contains "Monday"
    header_row = None
    for row in ws.iter_rows(min_col=2, max_col=2):
        if "monday" in _clean(row[0].value).lower():
            header_row = row[0].row
            break
    if header_row is None:
        raise ValueError("Could not find a row with 'Monday' in column B to use as the header row.")

    rows = {
        "daily_soup": find_row(LABELS["daily_soup"]),
        "soup_of_week": find_row(LABELS["soup_of_week"]),
        "sandwich_of_day": find_row(LABELS["sandwich_of_day"]),
        "lunch_entree": find_row(LABELS["lunch_entree"]),
        "appetizer": find_row(LABELS["appetizer"]),
        "dinner_entree_1": find_row(LABELS["dinner_entree"], occurrence=1),
        "dinner_entree_2": find_row(LABELS["dinner_entree"], occurrence=2),
        "dinner_dessert": find_row(LABELS["dinner_dessert"]),
    }

    days_out = []
    for day_name, col in zip(DAY_NAMES, DAY_COLS):
        header_text = _clean(ws[f"{col}{header_row}"].value)
        # header_text looks like "Monday   August 24" (no year) -> rebuild cleanly
        m = re.search(r"([A-Za-z]+)\s+([A-Za-z]+)\s+(\d{1,2})", header_text)
        if m:
            month_name, day_num = m.group(2), m.group(3)
        else:
            # fallback: just use the day name + whatever's left
            month_name, day_num = "", ""
        date_str = f"{day_name}, {month_name} {day_num}, {year}".upper() if month_name else day_name.upper()

        day_data = {"menu_date": date_str}
        for key, row_num in rows.items():
            day_data[key] = _clean(ws[f"{col}{row_num}"].value)
        days_out.append(day_data)

    return days_out


if __name__ == "__main__":
    import sys, json
    path = sys.argv[1]
    result = parse_weekly_menu(path, year=2026)
    print(json.dumps(result, indent=2))
